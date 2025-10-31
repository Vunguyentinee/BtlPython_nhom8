import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, StringVar
from datetime import datetime
import random
import cv2
import os
import subprocess,sys 
import threading
import mysql.connector
import hashlib
import unicodedata, re
from remove_person import (
    remove_from_encodings,
    delete_employee_in_db,
    delete_dataset_folder,
)
from capture_face import insert_employee,FaceCollector
from pathlib import Path

ROOT = Path(r"E:\BTL_Python")

DB_CONFIG = {
    "host": "127.0.0.1",
    "user": "root",
    "password": "21092005",
    "database": "dulieu_app",
    "port": 3306,
    "autocommit": True,
}

def db_conn():
    return mysql.connector.connect(**DB_CONFIG)
'''
def insert_employee_db(ma_nv, ten, phongban, chucvu, email, sdt):
    sql_check = "SELECT 1 FROM nhanvien WHERE ma_nv=%s LIMIT 1"
    sql_ins = """INSERT INTO nhanvien(ma_nv, ten, phongban, chucvu, email, sdt)
                 VALUES (%s,%s,%s,%s,%s,%s)"""
    with db_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql_check, (ma_nv,))
            if cur.fetchone():
                return False  # đã tồn tại
            cur.execute(sql_ins, (ma_nv, ten, phongban, chucvu, email, sdt))
    return True
'''
def get_employee_by_id(ma_nv):
    sql = "SELECT ten, phongban, chucvu, email, sdt FROM nhanvien WHERE ma_nv=%s"
    with db_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (ma_nv,))
            row = cur.fetchone()
            return row  # None nếu không có
        
def update_employee_db(ma_nv, ten, phongban, chucvu, email, sdt):
    sql = """UPDATE nhanvien
             SET ten=%s, phongban=%s, chucvu=%s, email=%s, sdt=%s
             WHERE ma_nv=%s"""
    with db_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (ten, phongban, chucvu, email, sdt, ma_nv))


def hash_pw(raw: str) -> str:
    # dùng SHA256 đơn giản (nếu muốn mạnh hơn có thể chuyển sang bcrypt)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()

def account_exists(username: str) -> bool:
    sql = "SELECT 1 FROM taikhoan WHERE ten_dang_nhap=%s LIMIT 1"
    with db_conn() as cn, cn.cursor() as cur:
        cur.execute(sql, (username,))
        return cur.fetchone() is not None
    
def employee_exists(ma_nv: str) -> bool:
    sql = "SELECT 1 FROM nhanvien WHERE ma_nv=%s LIMIT 1"
    with db_conn() as cn, cn.cursor() as cur:
        cur.execute(sql, (ma_nv,))
        return cur.fetchone() is not None
    
def insert_account(ten_that: str, ma_nv: str, username: str, raw_password: str):
    sql = """
      INSERT INTO taikhoan(ten_that, ma_nv, ten_dang_nhap, mat_khau)
      VALUES (%s, %s, %s, %s)
    """
    with db_conn() as cn, cn.cursor() as cur:
        cur.execute(sql, (ten_that, ma_nv, username, hash_pw(raw_password)))

def get_password_hash(username: str) -> str | None:
    sql = "SELECT mat_khau FROM taikhoan WHERE ten_dang_nhap=%s LIMIT 1"
    with db_conn() as cn, cn.cursor() as cur:
        cur.execute(sql, (username,))
        row = cur.fetchone()
        return row[0] if row else None
    
def update_password(username: str, new_raw_password: str) -> bool:
    sql = "UPDATE taikhoan SET mat_khau=%s WHERE ten_dang_nhap=%s"
    with db_conn() as cn, cn.cursor() as cur:
        cur.execute(sql, (hash_pw(new_raw_password), username))
        return cur.rowcount > 0

def safe_slug(s: str) -> str:
    # bỏ dấu -> ASCII
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    # chỉ giữ [A-Za-z0-9._-]
    s = re.sub(r'[^A-Za-z0-9._-]+', '_', s).strip('_')
    return s


# --------------------
# Config (soft blue)
# --------------------
PRIMARY = "#3b82f6"
PRIMARY_DARK = "#2563eb"
BG = "#f6f9ff"
CARD = "#ffffff"
SUBTEXT = "#50616a"

DATA_FILE = "admin_account.txt"   # ✅ Lưu tài khoản Admin (demo)



# ---------- App ----------
class LoginWindow(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Đăng nhập — Ứng dụng Chấm công nhận diện khuôn mặt")
        self.geometry("420x380")
        self.configure(bg=BG)
        self.resizable(False, False)
        self._build_ui()

    def _build_ui(self):
        frm = ttk.Frame(self, padding=18, style="Card.TFrame")
        frm.place(relx=0.5, rely=0.5, anchor="center")

        ttk.Label(frm, text="Ứng dụng Chấm công nhận diện khuôn mặt",
                  font=("Segoe UI Semibold", 12)).pack(pady=(0,10))

        ttk.Label(frm, text="Tên đăng nhập:", background=CARD).pack(anchor="w", pady=(6,0))
        self.username = ttk.Entry(frm, width=36)
        self.username.pack(pady=(0,6))

        ttk.Label(frm, text="Mật khẩu:", background=CARD).pack(anchor="w", pady=(6,0))
        self.password = ttk.Entry(frm, show="*", width=36)
        self.password.pack(pady=(0,6))

        hint = "Nếu chưa có tài khoản Admin → hãy nhấn Đăng ký"
        ttk.Label(frm, text=hint, font=("Segoe UI", 8),
                  foreground=SUBTEXT, background=CARD).pack(anchor="w", pady=6)

        btn_group = ttk.Frame(frm)
        btn_group.pack(fill="x", pady=6)

        ttk.Button(btn_group, text="Đăng nhập", command=self._on_login, style="Accent.TButton")\
            .pack(side="left", expand=True, fill="x", padx=(0,6))
        ttk.Button(btn_group, text="Đăng ký", command=self._register_user)\
            .pack(side="left", expand=True, fill="x", padx=(6,0))
        ttk.Button(frm, text="Đổi mật khẩu", command=self._open_change_password)\
            .pack(fill="x", pady=(0,6))

        ttk.Button(frm, text="Thoát", command=self.quit).pack(fill="x", pady=(6,0))

        # Style
        style = ttk.Style(self)
        try: style.theme_use("clam")
        except: pass
        style.configure("Card.TFrame", background=CARD)
        style.configure("Accent.TButton", background=PRIMARY, foreground="white", padding=8)
        style.map("Accent.TButton", background=[("active", PRIMARY_DARK)])

    def _read_account(self):
        # Nếu chưa có file hoặc file rỗng -> không có tài khoản
        if not os.path.exists(DATA_FILE) or os.path.getsize(DATA_FILE) == 0:
            return None, None

        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = f.read().strip().split("|")

                # đúng định dạng: fullname|emp_id|username|password
                if len(data) == 4:
                    return data[2], data[3]   # username, password

        except:
            pass

        return None, None



    def _on_login(self):
        user = self.username.get().strip()
        pwd = self.password.get().strip()

        if user == "admin" and pwd == "admin123":
            messagebox.showinfo("Thành công", f"Xin chào {user}!")
            Dashboard(self, role="Admin", username=user)
            self.withdraw()
            return

        stored_user, stored_pwd = self._read_account()

        if not user or not pwd:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập tên đăng nhập và mật khẩu.")
            return

        try:
            ph = get_password_hash(user)
        except Exception as e:
            messagebox.showerror("Lỗi MySQL", f"Không thể kiểm tra tài khoản:\n{e}")
            return

        if ph and hash_pw(pwd) == ph:
            messagebox.showinfo("Thành công", f"Xin chào {user}!")
            Dashboard(self, role="Admin", username=user)
            self.withdraw()
        else:
            messagebox.showerror("Lỗi đăng nhập", "Sai tên đăng nhập hoặc mật khẩu!")

    # ✅ giữ nguyên popup đăng ký tài khoản như bạn muốn
    # ✅ popup đăng ký tài khoản Admin (đã sửa thành 5 trường)
    def _register_user(self):
        dlg = tk.Toplevel(self)
        dlg.title("Đăng ký Admin")
        dlg.geometry("400x340")
        dlg.configure(bg=BG)
        dlg.grab_set()

        frame = ttk.Frame(dlg, padding=12, style="Card.TFrame")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Tạo tài khoản Admin", font=("Segoe UI Semibold", 12)).pack(pady=(0,5))

        ttk.Label(frame, text="Họ tên (tên thật):").pack(anchor="w")
        fullname = ttk.Entry(frame); fullname.pack(fill="x")

        ttk.Label(frame, text="Mã nhân viên:").pack(anchor="w", pady=(6,0))
        emp_id = ttk.Entry(frame); emp_id.pack(fill="x")

        ttk.Label(frame, text="Tên đăng nhập:").pack(anchor="w", pady=(6,0))
        username = ttk.Entry(frame); username.pack(fill="x")

        ttk.Label(frame, text="Mật khẩu:").pack(anchor="w", pady=(6,0))
        pwd = ttk.Entry(frame, show="*"); pwd.pack(fill="x")

        ttk.Label(frame, text="Nhập lại mật khẩu:").pack(anchor="w", pady=(6,0))
        pwd2 = ttk.Entry(frame, show="*"); pwd2.pack(fill="x")

        def save_account():
            ten_that = fullname.get().strip()
            ma_nv    = emp_id.get().strip()
            user     = username.get().strip()
            p1       = pwd.get().strip()
            p2       = pwd2.get().strip()

            if not ten_that or not ma_nv or not user or not p1 or not p2:
                messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ các trường.")
                return
            if p1 != p2:
                messagebox.showerror("Sai mật khẩu", "Mật khẩu nhập lại không trùng khớp!")
                return

            # (tuỳ) kiểm tra mã NV có tồn tại
            try:
                if not employee_exists(ma_nv):
                    if not messagebox.askyesno("Xác nhận",
                        "Mã nhân viên chưa tồn tại trong bảng NHANVIEN.\nBạn vẫn muốn tạo tài khoản chứ?"):
                        return
            except Exception as e:
                messagebox.showerror("Lỗi MySQL", f"Không kiểm tra được nhân viên:\n{e}")
                return

                # kiểm tra trùng username
            try:
                if account_exists(user):
                    messagebox.showerror("Trùng tài khoản", "Tên đăng nhập đã tồn tại, hãy chọn tên khác.")
                    return
                insert_account(ten_that, ma_nv, user, p1)
            except mysql.connector.IntegrityError as e:
                messagebox.showerror("Lỗi dữ liệu", f"Không thể lưu tài khoản:\n{e}")
                return
            except Exception as e:
                messagebox.showerror("Lỗi MySQL", f"Không thể lưu tài khoản:\n{e}")
                return

            messagebox.showinfo("Thành công", "Đăng ký tài khoản thành công! Hãy đăng nhập.")
            # điền sẵn username vào ô đăng nhập
            self.username.delete(0, tk.END)
            self.username.insert(0, user)
            self.password.focus()
            dlg.destroy()

        ttk.Button(frame, text="Lưu tài khoản", command=save_account, style="Accent.TButton").pack(pady=14)
    
    def _open_change_password(self):
        ChangePasswordWindow(self)

class ChangePasswordWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Đổi mật khẩu")
        self.geometry("380x260")
        self.configure(bg=BG)
        self.resizable(False, False)

        frame = ttk.Frame(self, padding=12, style="Card.TFrame")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Tên đăng nhập:").pack(anchor="w")
        self.username = ttk.Entry(frame)
        self.username.pack(fill="x")

        ttk.Label(frame, text="Mật khẩu cũ:").pack(anchor="w", pady=(6,0))
        self.old_pw = ttk.Entry(frame, show="*")
        self.old_pw.pack(fill="x")

        ttk.Label(frame, text="Mật khẩu mới:").pack(anchor="w", pady=(6,0))
        self.new_pw = ttk.Entry(frame, show="*")
        self.new_pw.pack(fill="x")

        ttk.Button(frame, text="Xác nhận", command=self._change_password, style="Accent.TButton").pack(pady=12)

    def _change_password(self):
        user   = self.username.get().strip()
        old_pw = self.old_pw.get().strip()
        new_pw = self.new_pw.get().strip()

        if not user or not old_pw or not new_pw:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ các trường.")
            return

        try:
            ph = get_password_hash(user)
        except Exception as e:
            messagebox.showerror("Lỗi MySQL", f"Không thể đọc tài khoản:\n{e}")
            return

        if not ph or hash_pw(old_pw) != ph:
            messagebox.showerror("Sai thông tin", "Tên đăng nhập hoặc mật khẩu cũ không đúng!")
            return

        try:
            ok = update_password(user, new_pw)
        except Exception as e:
            messagebox.showerror("Lỗi MySQL", f"Không thể cập nhật mật khẩu:\n{e}")
            return

        if ok:
            messagebox.showinfo("Thành công", "Đổi mật khẩu thành công!")
            self.destroy()
        else:
            messagebox.showwarning("Không thay đổi", "Không cập nhật được mật khẩu.")


class Dashboard(tk.Toplevel):
    def load_employees_from_db(self):
        """Đọc bảng nhanvien và đổ vào self.employee_list."""
        self.employee_list = []
        try:
            with db_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT ma_nv, ten, phongban, chucvu
                        FROM nhanvien
                        ORDER BY ma_nv ASC
                    """)
                    for ma, ten, pb, cv in cur.fetchall():
                        self.employee_list.append({
                            "id": ma or "",
                            "name": ten or "",
                            "dept": pb or "",
                            "role": cv or ""
                        })
        except Exception as e:
            messagebox.showerror("Lỗi MySQL", f"Không thể tải danh sách nhân viên:\n{e}")

    def __init__(self, master, role="Staff", username="user"):
        super().__init__(master)
        self.title("Ứng dụng Chấm công nhận diện khuôn mặt")
        self.geometry("1100x700")
        self.configure(bg=BG)
        self.role = role
        self.username = username
        self.employee_list = []
        self.tree_emp = None
        self._build_style()
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Header.TFrame", background=PRIMARY)
        style.configure("Header.TLabel", background=PRIMARY, foreground="white", font=("Segoe UI Semibold", 14))
        style.configure("TopMenu.TButton", background=PRIMARY, foreground="white", padding=8)
        style.configure("Card.TFrame", background=CARD, relief="flat")
        style.configure("Title.TLabel", background=CARD, font=("Segoe UI Semibold", 14))
        style.configure("Sub.TLabel", background=CARD, foreground=SUBTEXT)
        style.configure("AccentSmall.TButton", background=PRIMARY, foreground="white", padding=6)
        style.map("AccentSmall.TButton", background=[("active", PRIMARY_DARK)])
        style.configure("Treeview", background="white", fieldbackground="white", rowheight=28, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 10))
        style.map("Treeview", background=[("selected", "#dbeafe")], foreground=[("selected", "black")])

    def _build_ui(self):
        header = ttk.Frame(self, style="Header.TFrame", padding=10)
        header.pack(side="top", fill="x")
        ttk.Label(header, text="Ứng dụng Chấm công nhận diện khuôn mặt", style="Header.TLabel").pack(side="left", padx=12)
        ttk.Label(header, text=f"{self.role}: {self.username}", background=PRIMARY, foreground="white").pack(side="right", padx=12)

        menu_bar = ttk.Frame(self, padding=8, style="Card.TFrame")
        menu_bar.pack(side="top", fill="x", padx=12, pady=(12,6))

        self.menu_buttons = {}
        menu_items_admin = [("Trang chủ", self.show_home),
                            ("Nhân viên", self.show_employees),
                            ("Chấm công", self.show_attendance),
                            ("Báo cáo", self.show_reports),
                            ("Cấu hình", self.show_settings)]
        menu_items_staff = [("Chấm công", self.show_attendance),
                            ("Xin nghỉ phép", self.show_leave_request)]
        items = menu_items_admin if self.role == "Admin" else menu_items_staff

        for (label, cmd) in items:
            b = ttk.Button(menu_bar, text=label, command=cmd, style="AccentSmall.TButton")
            b.pack(side="left", padx=6)
            self.menu_buttons[label] = b

        right_frame = ttk.Frame(menu_bar, style="Card.TFrame")
        right_frame.pack(side="right")
        self.search_var = StringVar()
        self.search_entry = ttk.Entry(right_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side="left", padx=(0,6))
        self.search_entry.insert(0, "Tìm nhân viên…")
        self.search_entry.bind("<FocusIn>", lambda e: self._clear_placeholder())
        self.search_entry.bind("<KeyRelease>", lambda e: self.on_search())
        clear_btn = ttk.Button(right_frame, text="✖", command=self.clear_search)
        clear_btn.pack(side="left")

        self.content = ttk.Frame(self, style="Card.TFrame", padding=12)
        self.content.pack(fill="both", expand=True, padx=12, pady=(0,12))
        self.show_home()

    def _clear_placeholder(self):
        if self.search_var.get().strip().lower() == "tìm nhân viên…":
            self.search_var.set("")


    def clear_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    # ------- Pages -------
    def show_home(self):
        self.clear_content()
        card = ttk.Frame(self.content, style="Card.TFrame", padding=16)
        card.pack(fill="both", expand=True)
        ttk.Label(card, text="🏠 Trang chủ", style="Title.TLabel").pack(anchor="w")
        ttk.Label(card, text=f"Chào mừng {self.username} ({self.role})!", style="Sub.TLabel").pack(anchor="w", pady=(6,12))
        stats_frame = ttk.Frame(card, style="Card.TFrame")
        stats_frame.pack(fill="x", pady=6)
        for title, val in [("Nhân viên", str(len(self.employee_list))), ("Chấm công hôm nay", "9"), ("Vắng", "3")]:
            box = ttk.Frame(stats_frame, style="Card.TFrame", padding=12)
            box.pack(side="left", padx=8)
            ttk.Label(box, text=title, font=("Segoe UI", 10)).pack(anchor="w")
            ttk.Label(box, text=val, font=("Segoe UI Semibold", 12)).pack(anchor="w")

    def show_employees(self):
        self.clear_content()
        card = ttk.Frame(self.content, style="Card.TFrame", padding=12)
        card.pack(fill="both", expand=True)

        ttk.Label(card, text="👨‍💼 Quản lý nhân viên", style="Title.TLabel").pack(anchor="w")

        # Topbar
        topbar = ttk.Frame(card, style="Card.TFrame")
        topbar.pack(fill="x", pady=(6,8))

        if self.role == "Admin":
            ttk.Button(topbar, text="➕ Thêm nhân viên", command=self.add_employee).pack(side="right", padx=6)
            ttk.Button(topbar, text="🗑 Xóa nhân viên", command=self.delete_employee).pack(side="right", padx=6)

        # Nút làm mới từ DB
        ttk.Button(topbar, text="🔄 Làm mới (DB)", command=self.refresh_employees).pack(side="left")

        cols = ("Mã NV", "Họ tên", "Phòng ban", "Chức vụ")
        self.tree_emp = ttk.Treeview(card, columns=cols, show="headings")
        for c in cols:
            self.tree_emp.heading(c, text=c)
            self.tree_emp.column(c, width=200 if c == "Họ tên" else 140, anchor="w")
        self.tree_emp.pack(fill="both", expand=True, pady=(6,0))

        # Lần đầu: lấy từ DB
        self.refresh_employees()

    def refresh_employees(self):
        """Đổ self.employee_list ra Treeview"""
        if not self.tree_emp:
            return
        # nếu muốn luôn đọc mới từ DB mỗi lần refresh:
        self.load_employees_from_db()
        self.tree_emp.delete(*self.tree_emp.get_children())
        for emp in self.employee_list:
            self.tree_emp.insert("", "end",
                values=(emp["id"], emp["name"], emp["dept"], emp["role"]))

    def show_attendance(self):
        self.clear_content()
        card = ttk.Frame(self.content, style="Card.TFrame", padding=12)
        card.pack(fill="both", expand=True)

        ttk.Label(card, text="🕒 Chấm công", style="Title.TLabel").pack(anchor="w")

        topbar = ttk.Frame(card, style="Card.TFrame")
        topbar.pack(fill="x", pady=(6,8))

        # Nút làm mới (gọi refresh_attendance)
        ttk.Button(topbar, text="🔄 Làm mới", command=self.refresh_attendance)\
        .pack(side="left", padx=6)

        # nút mở camera nhận diện
        ttk.Button(topbar, text="Check-in FaceID", command=self.start_faceid).pack(side="right", padx=6)

        # Bảng hiển thị: Thời gian vào / Mã NV / Họ tên / Ghi chú
        cols = ("Thời gian vào", "Mã NV", "Họ tên", "Ghi chú")
        self.tree_att = ttk.Treeview(card, columns=cols, show="headings")
        for c in cols:
            self.tree_att.heading(c, text=c)
            self.tree_att.column(c, width=200 if c == "Họ tên" else 160 if c == "Thời gian vào" else 120, anchor="w")
        self.tree_att.pack(fill="both", expand=True, pady=(6,0))

        self.refresh_attendance()

    def refresh_attendance(self):
        """Lấy danh sách chấm công hôm nay và đổ vào TreeView."""
        try:
            with db_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                    SELECT 
                        CONCAT(DATE_FORMAT(ngay, '%Y-%m-%d'), ' ', 
                               IFNULL(DATE_FORMAT(gio_checkin, '%H:%i:%s'), ''))  AS thoi_gian_vao,
                        ma_nv,
                        ten_nv,
                        IFNULL(ghichu,'') AS ghichu
                    FROM chamcong
                    WHERE ngay = CURDATE()
                    ORDER BY gio_checkin
                """)
                    rows = cur.fetchall()
        except Exception as e:
            messagebox.showerror("Lỗi DB", f"Không thể tải chấm công hôm nay:\n{e}")
            return

    # clear & fill
        for it in self.tree_att.get_children():
            self.tree_att.delete(it)
        for r in rows:
            self.tree_att.insert("", "end", values=r)

    def start_faceid(self):
        """Chạy script nhận diện; khi tắt camera sẽ refresh bảng."""
        def run():
            try:
                # đảm bảo dùng đúng Python trong venv hiện tại
                subprocess.run([sys.executable, str(ROOT / "recognize_checkin_out.py")], check=False)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Lỗi", f"Không chạy được nhận diện:\n{e}"))
                return
            # sau khi user nhấn q/ESC đóng camera -> refresh bảng
            self.after(0, self.refresh_attendance)

        threading.Thread(target=run, daemon=True).start()

    def show_reports(self):
        self.clear_content()
        card = ttk.Frame(self.content, style="Card.TFrame", padding=16)
        card.pack(fill="both", expand=True)
        ttk.Label(card, text="📊 Báo cáo ", style="Title.TLabel").pack(anchor="w")
        ttk.Label(card, text="Tính năng xuất báo cáo sẽ do backend thực hiện; hiện đây là khung demo.", style="Sub.TLabel").pack(anchor="w", pady=8)
        ttk.Button(card, text="Xuất Excel ", command=self.export_excel).pack(pady=8)

    def show_settings(self):
        self.clear_content()
        card = ttk.Frame(self.content, style="Card.TFrame", padding=16)
        card.pack(fill="both", expand=True)
        ttk.Label(card, text="⚙️ Cấu hình hệ thống (demo)", style="Title.TLabel").pack(anchor="w")
        ttk.Label(card, text="Các thiết lập (tạm) sẽ do backend quản lý.", style="Sub.TLabel").pack(anchor="w", pady=8)

    def show_leave_request(self):
        self.clear_content()
        card = ttk.Frame(self.content, style="Card.TFrame", padding=16)
        card.pack(fill="both", expand=True)
        ttk.Label(card, text="📝 Xin nghỉ phép", style="Title.TLabel").pack(anchor="w")
        ttk.Label(card, text="Lý do:", style="Sub.TLabel").pack(anchor="w", pady=(6,0))
        reason = ttk.Entry(card, width=60); reason.pack(anchor="w", pady=4)
        ttk.Label(card, text="Từ ngày (YYYY-MM-DD):", style="Sub.TLabel").pack(anchor="w", pady=(6,0))
        start = ttk.Entry(card, width=20); start.pack(anchor="w", pady=4)
        ttk.Label(card, text="Đến ngày (YYYY-MM-DD):", style="Sub.TLabel").pack(anchor="w", pady=(6,0))
        end = ttk.Entry(card, width=20); end.pack(anchor="w", pady=4)
        def send_request():
            if not reason.get().strip() or not start.get().strip() or not end.get().strip():
                messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ thông tin!")
                return
            messagebox.showinfo("Đã gửi", "Yêu cầu nghỉ phép đã được gửi (demo).")
        ttk.Button(card, text="Gửi yêu cầu", command=send_request).pack(pady=10)

    def _start_capture_and_encode(self, ten: str, ma: str):
        folder_name = safe_slug(f"{ten}_{ma}")
        save_root = str(ROOT / "dataset")

        def run_collect():
            ok = False
            err = None
            try:
                ok = FaceCollector(save_root=save_root).collect(folder_name)
            except Exception as e:
                err = e

            def done():
                if err:
                    messagebox.showerror("Quét mặt", f"Lỗi khi thu ảnh:\n{err}")
                elif ok:
                    messagebox.showinfo("Quét mặt", f"✔ Đã thu ảnh cho {ten} ({ma}).\nThư mục: {save_root}\\{folder_name}")
                    # gọi encode_sync ngay sau khi chụp xong
                    try:
                        subprocess.run([sys.executable, str(ROOT / "encode_sync.py")], check=False)
                    except Exception as enc_ex:
                        messagebox.showwarning("Encode", f"Không chạy được encode_sync.py:\n{enc_ex}")
                else:
                    messagebox.showwarning("Quét mặt", "Chưa thu được ảnh nào.")
            self.after(0, done)

        threading.Thread(target=run_collect, daemon=True).start()
    # ------- small features -------
    # -------------------------------------
    # ✅ THÊM NHÂN VIÊN + QUÉT KHUÔN MẶT
    # -------------------------------------


    def add_employee(self):
        dlg = tk.Toplevel(self)
        dlg.title("Thêm nhân viên mới")
        dlg.geometry("420x350")
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=12, style="Card.TFrame")
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Thêm nhân viên mới", style="Title.TLabel").pack(pady=6)

        ttk.Label(frm, text="Mã nhân viên").pack(anchor="w")
        eid = ttk.Entry(frm); eid.pack(fill="x")

        ttk.Label(frm, text="Họ tên").pack(anchor="w")
        name = ttk.Entry(frm); name.pack(fill="x")

        ttk.Label(frm, text="Phòng ban").pack(anchor="w")
        dept = ttk.Entry(frm); dept.pack(fill="x")

        ttk.Label(frm, text="Chức vụ").pack(anchor="w")
        role = ttk.Entry(frm); role.pack(fill="x")

        ttk.Label(frm, text="Email liên hệ").pack(anchor="w")
        email = ttk.Entry(frm); email.pack(fill="x")

        ttk.Label(frm, text="Số điện thoại liên hệ").pack(anchor="w")
        phone = ttk.Entry(frm); phone.pack(fill="x")

        def submit():
            ma = eid.get().strip()
            ten = name.get().strip()
            pb  = dept.get().strip()
            cv  = role.get().strip()
            em  = email.get().strip()
            sdt = phone.get().strip()

            if not ma or not ten:
                messagebox.showwarning("Thiếu thông tin", "Cần nhập tối thiểu: Mã nhân viên và Họ tên.")
                return

            try:
                with db_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT 1 FROM nhanvien WHERE ma_nv=%s LIMIT 1", (ma,))
                        existed = cur.fetchone() is not None
            except Exception as e:
                messagebox.showerror("Lỗi MySQL", f"Không thể kiểm tra nhân viên:\n{e}")
                return
            
            if existed:
                # Hỏi hướng xử lý
                choice = messagebox.askyesnocancel(
                    "Nhân viên đã tồn tại",
                    f"Mã {ma} đã có.\n"
                    f"[Yes] Quét lại khuôn mặt & encode thêm\n"
                    f"[No]  Cập nhật thông tin rồi quét\n"
                    f"[Cancel] Huỷ"
                )
                if choice is None:
                    return
                if choice:  # YES -> chỉ quét
                    self._start_capture_and_encode(ten, ma)
                    dlg.destroy()
                    return
                else:       # NO -> cập nhật + quét
                    try:
                        with db_conn() as conn:
                            with conn.cursor() as cur:
                                cur.execute("""
                                    UPDATE nhanvien
                                    SET ten=%s, phongban=%s, chucvu=%s, email=%s, sdt=%s
                                    WHERE ma_nv=%s
                                    """, (ten, pb, cv, em, sdt, ma))
                        messagebox.showinfo("Cập nhật", f"Đã cập nhật thông tin {ma}.")
                    except Exception as e:
                        messagebox.showerror("Lỗi MySQL", f"Không thể cập nhật nhân viên:\n{e}")
                        return
                    # reload list từ DB để bảng UI đúng thực tế
                    self.refresh_employees()
                    self._start_capture_and_encode(ten, ma)
                    dlg.destroy()
                    return
            else:
                # Chèn mới + refresh + quét
                try:
                    insert_employee(ma, ten, pb, cv, em, sdt)
                except Exception as e:
                    messagebox.showerror("Lỗi MySQL", f"Lưu nhân viên thất bại:\n{e}")
                    return
                self.refresh_employees()
                self._start_capture_and_encode(ten, ma)
                dlg.destroy()
                return

        ttk.Button(frm, text="Thêm & Quét mặt", command=submit).pack(pady=10)
    def delete_employee(self):
        dlg = tk.Toplevel(self)
        dlg.title("Xóa nhân viên")
        dlg.geometry("350x240")
        dlg.configure(bg=BG)
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=12, style="Card.TFrame")
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Xóa nhân viên", style="Title.TLabel").pack(pady=6)

        ttk.Label(frm, text="Mã nhân viên:").pack(anchor="w")
        emp_id = ttk.Entry(frm)
        emp_id.pack(fill="x", pady=(0,6))

        ttk.Label(frm, text="Họ tên nhân viên:").pack(anchor="w")
        emp_name = ttk.Entry(frm)
        emp_name.pack(fill="x")

        def confirm_delete():
            id_val = emp_id.get().strip()
            name_val = emp_name.get().strip()

            if not id_val or not name_val:
                messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ Mã nhân viên và Họ tên.")
                return

# 1) Xóa khỏi danh sách UI
            found = False
            for emp in list(self.employee_list):
                if emp["id"].lower() == id_val.lower() and emp["name"].lower() == name_val.lower():
                    self.employee_list.remove(emp)
                    found = True
                    break
            if not found:
                messagebox.showerror("Không tìm thấy", "Không tồn tại nhân viên này trong danh sách hiển thị.")
                return

        # 2) Xóa khỏi TreeView
            for item in self.tree_emp.get_children():
                values = self.tree_emp.item(item, "values")
                if values and values[0].lower() == id_val.lower() and values[1].lower() == name_val.lower():
                    self.tree_emp.delete(item)

        # 3) Xóa ENCODINGS + DB + DATASET (dùng module remove_person.py)
            try:
                label_full = f"{name_val}_{id_val}"        # Tên_MãNV đúng chuẩn encodings
            # encodings: thử xóa theo full label trước, nếu không có thì theo mã
                enc_removed = remove_from_encodings(label_full)
                if enc_removed == 0:
                    remove_from_encodings(id_val)

            # DB: hàm của bạn chấp nhận cả "Tên_MãNV" hoặc chỉ "MãNV"
                delete_employee_in_db(label_full)

            # dataset folder: xóa chính xác thư mục ảnh
                delete_dataset_folder(label_full)

                messagebox.showinfo("Thành công", f"Đã xóa '{name_val} ({id_val})' khỏi UI, encodings, DB và dataset.")
                dlg.destroy()
            except Exception as e:
                messagebox.showerror("Lỗi khi xóa", str(e))

        ttk.Button(frm, text="Xóa", command=confirm_delete, style="AccentSmall.TButton").pack(pady=12)



    def capture_faces_and_add_employee(self, emp_id=None, emp_name=None):
        print(">>> [UI] Đã gọi hàm quét mặt — chờ backend xử lý...")
        print(f">>> Employee ID: {emp_id}, Name: {emp_name}")

        # 📌 Tổng kết:
        #  - UI đã gọi đúng hàm
        #  - Backend sau này sẽ đưa code capture vào đây
        messagebox.showinfo(
            "Quét mặt",
            f"📷 Đang quét khuôn mặt cho nhân viên:\n\n"
            f"➡️ {emp_name} ({emp_id})\n\n"
            f"(UI đã gọi hàm — Backend xử lý tiếp)"
        )



    def fake_checkin(self):
        # simple checkin demo (no anti-spoofing here)
        messagebox.showinfo("Check-in", "Nhận diện khuôn mặt thành công (demo). Dữ liệu đã gửi lên server (demo).")

    def export_excel(self):
        """
        Xuất bảng chấm công ra Excel từ MySQL (bảng chamcong)
        Có popup chọn khoảng ngày và tự mở file sau khi xuất.
        """
        from tkinter import Toplevel, ttk
        from datetime import datetime, date
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import subprocess

        # --- Popup chọn khoảng thời gian ---
        dlg = Toplevel(self)
        dlg.title("Chọn khoảng thời gian xuất Excel")
        dlg.geometry("360x210")
        dlg.configure(bg="#f6f9ff")
        dlg.grab_set()

        frame = ttk.Frame(dlg, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Từ ngày (YYYY-MM-DD):").pack(anchor="w", pady=(4, 0))
        entry_from = ttk.Entry(frame, width=25)
        entry_from.pack(anchor="w", pady=(0, 6))

        ttk.Label(frame, text="Đến ngày (YYYY-MM-DD):").pack(anchor="w", pady=(4, 0))
        entry_to = ttk.Entry(frame, width=25)
        entry_to.pack(anchor="w", pady=(0, 8))

        def _parse_date(s: str):
            s = s.strip()
            if not s:
                return None
            try:
                return datetime.strptime(s, "%Y-%m-%d").date()
            except ValueError:
                return None
            
        def run_export():
            # Nếu để trống -> mặc định hôm nay
            d_from = _parse_date(entry_from.get()) or date.today()
            d_to   = _parse_date(entry_to.get())   or date.today()

            if d_from > d_to:
                messagebox.showerror("Sai khoảng ngày", "Từ ngày phải nhỏ hơn hoặc bằng Đến ngày.")
                return

            try:
                with db_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("""
                            SELECT ngay,
                                TIME_FORMAT(gio_checkin,  '%%H:%%i:%%s') AS cin,
                                TIME_FORMAT(gio_checkout, '%%H:%%i:%%s') AS cout,
                                ma_nv, ten_nv, IFNULL(ghichu,'') AS ghichu
                            FROM chamcong
                            WHERE ngay BETWEEN %s AND %s
                            ORDER BY ngay DESC, ma_nv ASC
                        """, (d_from, d_to))  # ✅ luôn đủ 2 tham số
                        rows = cur.fetchall()
            except Exception as e:
                messagebox.showerror("Lỗi MySQL", f"Không thể đọc dữ liệu chấm công:\n{e}")
                return

            if not rows:
                messagebox.showinfo("Xuất Excel", "Không có dữ liệu trong khoảng ngày đã chọn.")
                return

            # ---- Ghi Excel ----
            wb = Workbook(); ws = wb.active; ws.title = "ChamCong"
            headers = ["Ngày", "Giờ check-in", "Giờ check-out", "Mã NV", "Họ tên", "Ghi chú", "Trạng thái"]
            ws.append(headers)

            for ngay, cin, cout, ma, ten, note in rows:
                ngay_str = ngay.strftime("%Y-%m-%d") if hasattr(ngay, "strftime") else str(ngay)
                status = "valid" if (note or "").strip().upper() == "OK" else ("late" if (note or "").lower().startswith("muộn") else "")
                ws.append([ngay_str, cin or "", cout or "", ma or "", ten or "", note or "", status])

            header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center", vertical="center")
            thin = Border(
                left=Side(style="thin", color="999999"), right=Side(style="thin", color="999999"),
                top=Side(style="thin", color="999999"),  bottom=Side(style="thin", color="999999"),
            )
            for c in ws[1]:
                c.fill = header_fill; c.font = header_font; c.alignment = header_align
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for c in row:
                    c.border = thin; c.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                maxlen = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(maxlen + 2, 40)

            out_dir = (ROOT / "reports"); out_dir.mkdir(exist_ok=True)
            out_path = out_dir / f"ChamCong_{d_from}_{d_to}_{datetime.now().strftime('%H%M%S')}.xlsx"
            try:
                wb.save(out_path)
            except Exception as e:
                messagebox.showerror("Lỗi xuất file", f"Lưu file thất bại:\n{e}")
                return

            if messagebox.askyesno("Xuất Excel", f"✅ Đã xuất: {out_path}\n\nMở file ngay?"):
                try:
                    os.startfile(out_path)
                except Exception:
                    subprocess.Popen(["start", "", str(out_path)], shell=True)
            dlg.destroy()

        ttk.Button(frame, text="📤 Xuất Excel", command=run_export).pack(pady=(10, 6))
        ttk.Button(frame, text="❌ Hủy", command=dlg.destroy).pack()


    def on_search(self):
        kw = self.search_var.get().strip().lower()
        if not self.tree_emp:
            return
        # if empty show all
        if not kw or kw == "tìm nhân viên…":
            # reset to the demo list
            self.tree_emp.delete(*self.tree_emp.get_children())
            for emp in self.employee_list:
                self.tree_emp.insert("", "end", values=(emp["id"], emp["name"], emp["dept"], emp["role"]))
            return
        # filter
        filtered = [e for e in self.employee_list if kw in e["id"].lower() or kw in e["name"].lower()]
        self.tree_emp.delete(*self.tree_emp.get_children())
        for emp in filtered:
            self.tree_emp.insert("", "end", values=(emp["id"], emp["name"], emp["dept"], emp["role"]))

    def clear_search(self):
        self.search_var.set("")
        self.on_search()

    def _on_close(self):
        # when dashboard closed, show login window back
        self.master.deiconify()
        self.destroy()


if __name__ == "__main__":
    app = LoginWindow()
    app.mainloop()
